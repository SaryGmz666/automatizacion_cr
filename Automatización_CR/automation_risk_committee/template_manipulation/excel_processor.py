import attr
import re

from ..config import (
    PATH_TEMPLATES,
    pattern_decryptor,
    transform_date_string,
)
from openpyxl import load_workbook


@attr.s(slots=True)
class ExcelTemplateProcessor:
    name_template: str = attr.ib()
    required_dates: dict[str, str] = attr.ib()

    def open_excel(self):
        path = f"{PATH_TEMPLATES}{self.name_template}.xlsx"

        try:
            wb = load_workbook(path, data_only=False)
        except Exception as e:
            raise RuntimeError(
                f"No fue posible abrir la plantilla '{self.name_template}'. "
                "Verifique que el archivo exista y no esté corrupto."
            ) from e

        # Aqui cambiamos los encabezados de las tablas.
        self.replace_period_headers(wb)

        return wb

    def replace_period_headers(self, wb):
        """Sustituye los encabezados de periodo de la plantilla (ej. 'dic_comparative')
        por el formato necesario.
        """
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:

                    if (
                        isinstance(cell.value, str)
                        and cell.value in self.required_dates
                    ):
                        cell.value = transform_date_string(
                            self.required_dates[cell.value]
                        )

    def extract_ref_cells(self, wb) -> tuple[list, list]:

        # NOTE Patrón buscado en las celdas de la plantilla
        REF_REGEX = re.compile(r"ref\(", re.IGNORECASE)

        target_cells = []
        extracted_refs = []

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and REF_REGEX.search(cell.value):
                        target_cells.append(cell)
                        extracted_refs.extend(pattern_decryptor(cell.value))

        return target_cells, extracted_refs

    def extract_keys_and_periods(self, extracted_refs) -> tuple[list, set]:
        board_keys, periods = zip(*extracted_refs)

        return (
            list(set(board_keys)),
            set(periods),
        )

    def validate_required_dates(self, periods_from_template: set):
        input_periods = set(self.required_dates.keys())

        if periods_from_template != input_periods:
            missing = periods_from_template - input_periods
            extra = input_periods - periods_from_template

            msg = "Inconsistencia en los periodos requeridos:\n"

            if missing:
                msg += (
                    f" - Periodos usados en la plantilla pero no definidos: {missing}\n"
                )
            if extra:
                msg += f" - Periodos definidos pero no utilizados en la plantilla: {extra}\n"

            raise ValueError(msg)

    def run(self):
        wb = self.open_excel()

        target_cells, extracted_refs = self.extract_ref_cells(wb)
        necessary_keys, periods = self.extract_keys_and_periods(extracted_refs)

        self.validate_required_dates(periods)

        return {
            "workbook": wb,
            "target_cells": target_cells,
            "necessary_keys": necessary_keys,
            "required_dates": self.required_dates,
        }
